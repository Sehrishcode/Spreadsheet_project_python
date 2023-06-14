import openpyxl as xl
from openpyxl.chart import BarChart, Reference
def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]
    # print(cell.value)
    # print(sheet.max_row)
    # print(sheet.max_column)
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value *0.9
        print(corrected_price)
        correcte_price_cell = sheet.cell(row,4)
        correcte_price_cell.value = corrected_price
    values = Reference(sheet, 
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)
process_workbook("Python-Tutorial-Supplementary-Materials/Python Tutorial Supplementary Materials/transactions.xlsx")